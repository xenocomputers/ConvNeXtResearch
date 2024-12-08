'''
File name: extractData.py
Author(s): Aarohan Mishra
'''
from openpyxl import Workbook
import sys

def main():
    # Check to see whether or not the correct number of arguments have been passed in
    if len(sys.argv) != 3:
        print("Usage:")
        print("python3 extractData.py <dataPath> <savePath>")
        print("dataPath: Path to log.txt file")
        print("savePath: Path to file which should be saved (eg. exampleDir/example.xlsx)")
        return

    dataPath= sys.argv[1]
    savePath= sys.argv[2]
    # Define column names for spreadsheet
    columns = ["train_lr", "train_min_lr", "train_loss", "train_weight_decay", "test_loss", "test_acc1", "test_acc5", "epoch", "n_parameters", "test_loss_ema", "test_acc1_ema", "test_acc5_ema"]

    wb= Workbook()  # Define Workbook

    ws= wb.active   # Define current worksheet

    ws.title= "Results" # Set current worksheet name

    # Set the column names in cells in Worksheet
    for i in range(len(columns)):
        c= ws.cell(row= 1, column= i+1)
        c.value = columns[i]

    # Read all of the lines in the log.txt file
    with open(dataPath) as f:
        lines= f.readlines()

    # Loop through all the lines in the log.txt file and write the values to cells in the current worksheet
    for i in range(len(lines)):
        entry= lines[i]
        entry= entry[1:len(entry)-2]
        vals = entry.split(',')
        for j in range(len(vals)):
            c= ws.cell(row = i+2, column= j + 1)
            c.value = vals[j].split(':')[1]

    wb.save(savePath)   # Save the current workbook to the savePath entered
    print("Completed")  # Print completion message


if __name__ == '__main__':
    main()